##################################################################################################################################################################
## Libraries
import os
import sys
import pandas as pd
import logging
import datetime
import itertools
from pathlib import Path
from importlib import reload

import arcpy
from arcgis.gis import GIS

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))

import  src.classes.DataCatalog as dc 
from src.functions import meta
from src.constants.paths import LOG_DIR
from src.constants.values import DF_COLUMNS, SERVICE_ITEM_LOOKUP, SHEET_NAME

#################################################################################################################################################################################################################
## Globals
DATETIME_STR = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
#################################################################################################################################################################################################################
## Logging
reload(logging)
log_file =Path(LOG_DIR, "CompareMetadata",f"CompareMetadata_{DATETIME_STR}.log")

logging.getLogger().disabled = True
logging.getLogger("arcgis.gis._impl._portalpy").setLevel(logging.WARNING)
logging.getLogger("urllib3.connectionpool").setLevel(logging.WARNING)
logging.getLogger("requests_oauthlib.oauth2_session").setLevel(logging.WARNING)

logger = logging.getLogger(__name__)
file_handler = logging.FileHandler(log_file, mode='w')
formatter = logging.Formatter('%(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)
logger.setLevel(logging.INFO)
##################################################################################################################################################################
## Main
def main(gis_conn:GIS, gdb_path:Path, catalog_path:Path, output_excel:Path, text_type:str, web_app_categories:list=None, include_exclude:str=None)->None:
    DF_DICTIONARY = {}
    DF_LIST = []

    EXCEL_COLS = {}

    logger.info(f"Starting: {datetime.datetime.now()}")
    logger.info(f"Run by: {os.getlogin()}")
    logger.info(f"Run on: {datetime.datetime.now().strftime('%d/%m/%Y, %H:%M:%S')}")
    logger.info(f"Web App Categories: {web_app_categories}")
    logger.info(f"Output Excel Path: {output_excel}")
    logger.info(f"Output Text Type: {text_type}")
    


    catalog_df = pd.read_excel(io=catalog_path, sheet_name=SHEET_NAME,header=0,names=None)

    catalog_rows = dc.getCatalogRows(catalog_df, web_app_categories, include_exclude)

    logger.info(f"Checking Metadata...")

    for index, row in catalog_rows:
        row_obj = dc.DataCatalogRow(row, index, gdb_path, gis_conn)

        logger.info(row_obj.table_name)
        arcpy.AddMessage(row_obj.table_name)

        out_dictionary= meta.getMetadata(row_obj=row_obj, md_items=list(SERVICE_ITEM_LOOKUP.keys()), text_type=text_type)

        DF_DICTIONARY[row_obj.table_name] = out_dictionary
        
    logger.info(f"Creating DataFrame... {datetime.datetime.now()}")
    df = pd.DataFrame.from_dict(DF_DICTIONARY, "index", columns=DF_COLUMNS)


    ## I sort the Pandas DataFrame for ease of reading. This can be commented out to keep the original order of the Data Catalog.
    df = df.sort_index(axis=0)

    logger.info(f"Exporting DataFrame...")
    df.to_excel(excel_writer=output_excel, sheet_name="MetadataCompare",columns=DF_COLUMNS,index_label="Table Name")

    ########## Excel Workbook ##########
    ## This section deals with the formating of the Excel Workbook
    logger.info(f"Formatting Excel Report...{datetime.datetime.now()}")

    wb = openpyxl.load_workbook(output_excel) ## Creates the Workbook Object

    sheet_name = wb.sheetnames[0] ## Creates a list of strings of the sheet names in the workbook

    ws = wb[sheet_name] ## Creates the Work Sheet object

    ## this conditional checks if the table already exists. If the table exists it is deleted and a new table is created.
    if sheet_name in ws.tables:
        del ws.tables[sheet_name]

    table_dimensions = ws.calculate_dimension() ## grabs the dimensions. these are used when creating the Table Object

    tbl = Table(displayName="MetadataCompare", ref=table_dimensions) ## Creates the Table object using the grabbed dimensions

    style = TableStyleInfo(name=f"TableStyleMedium{2}",showFirstColumn=True, showLastColumn=False, showRowStripes=True, showColumnStripes=False) ## Formats the table. This is where the count comes in.

    tbl.tableStyleInfo = style ## Applies the style info to the Table Object
    ws.add_table(tbl) ## Adds the Table Object to the Worksheet object

    exist_b_cols = ["B", "C", ]

    title_cols = ["D", "E"]
    title_b_col="F"

    description_cols = [ "G", "H"]
    description_b_col = "I"

    summary_cols = ["J", "K"]
    summary_b_col = "L"

    tags_cols = ["M", "N"]
    tags_b_col = "O"

    credits_cols =  ["P", "Q"]
    credits_b_col = "R"

    access_cols = ["S", "T"]
    access_b_col = "U"


    ## the below deals with the Column Width and Wrapping
    ws.column_dimensions["A"].width = 35

    for i in (exist_b_cols + [title_b_col, description_b_col, summary_b_col, tags_b_col, credits_b_col, access_b_col]):
        ws.column_dimensions[i].width = 20

    for i in list(itertools.chain(title_cols, description_cols, summary_cols, tags_cols, credits_cols, access_cols)):
        ws.column_dimensions[i].width = 75

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)




    wb.save(output_excel)
    del wb

    arcpy.AddMessage(f"Excel Report Has Been Exported to:\n{output_excel}")
    arcpy.AddMessage(f"Opening Excel Report...")
    try:
        os.startfile(output_excel)
    except Exception as t:
        arcpy.AddWarning(f"Failed to Launch Excel")

    logger.info(f"Ending: {datetime.datetime.now()}")
