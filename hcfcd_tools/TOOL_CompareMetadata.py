##################################################################################################################################################################
## Libraries
import os
import re
import sys
import pandas as pd
from pandas import DataFrame
import logging
import datetime
import itertools
from pathlib import Path
from bs4 import BeautifulSoup
from bs4.element import Tag

import arcpy
from arcpy import metadata as md
from arcgis.gis import GIS

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

if str(Path(__file__).resolve().parents[1]) not in sys.path:
    sys.path.insert(0,str(Path(__file__).resolve().parents[1]))
from hcfcd_classes.DataCatalog import DataCatalogRow
from hcfcd_constants.values import ROOT_DIR


#################################################################################################################################################################################################################
## Logging ## Don't Change

log_file =Path(ROOT_DIR, "logs", "CompareMetadata","CompareMetadata.log")

logging.basicConfig(filename=log_file, filemode="w",format='%(name)s - %(levelname)s - %(message)s', level=logging.INFO)

logging.getLogger("arcgis.gis._impl._portalpy").setLevel(logging.WARNING)
logging.getLogger("urllib3.connectionpool").setLevel(logging.WARNING)
logging.getLogger("requests_oauthlib.oauth2_session").setLevel(logging.WARNING)

logger = logging.getLogger(__name__)
##################################################################################################################################################################
## Input Parameters
master_gdb_name = "Master"
##################################################################################################################################################################
## Constants
DF_DICTIONARY = {}
DF_LIST = []

DF_COLUMNS = ["Master - Exist", 
              "Spatial - Exist", 
              "Service - Exist", 
              "Master - title",
              "Spatial - title",
              "Service - title",
              "title - Match", 
              "Master - description",
              "Spatial - description",
              "Service - description",
              "description - Match", 
              "Master - summary",
              "Spatial - summary",
              "Service - summary",
              "summary - Match", 
              "Master - tags",
              "Spatial - tags",
              "Service - tags",
              "tags - Match",
              "Master - credits",
              "Spatial - credits",
              "Service - credits",
              "credits - Match", 
              "Master - accessConstraints",
              "Spatial - accessConstraints",
              "Service - accessConstraints",
              "accessConstraints - Match"
              ]

EXCEL_COLS = {}
SERVICE_ITEM_LOOKUP = {"title":"title", 
                       "description":"description", 
                       "summary":"snippet", 
                       "tags":"tags", 
                       "credits":"accessInformation", 
                       "accessConstraints":"licenseInfo"}
##################################################################################################################################################################
## Functions

def _cleanCheckText(text):
    notags_text = re.sub(r'<.*?>', '', text)
    lower_text = notags_text.lower()
    
    #nofollow_text = lower_text.replace(" rel='nofollow ugc'", "")
    standard_quotes = lower_text.replace('"',"'")
    text_strip = standard_quotes.strip()
    clean_text = text_strip
    
    return clean_text

def getMetadata(row_obj:DataCatalogRow, md_items:list, text_type:str)->dict:

    out_dictionary = {}

    spatial_md_obj =md.Metadata(row_obj.getFilePath("Spatial"))
    master_md_obj = md.Metadata(row_obj.getFilePath("Master"))
    service_obj = row_obj.getServiceObject()
    
    master_exist = row_obj.master_exist
    spatial_exist = row_obj.spatial_exist
    service_exist = row_obj.service_exist

    out_dictionary[f"Master - Exist"] = master_exist
    out_dictionary[f"Spatial - Exist"] = spatial_exist
    out_dictionary[f"Service - Exist"] = service_exist

    for md_item in md_items:

        if not master_exist:
            master_md_attr = "Dataset Doesn't Exist"
        else:
            if hasattr(master_md_obj, md_item):
                master_md_attr = _formatMdItems(getattr(master_md_obj, md_item),md_item, text_type) if getattr(master_md_obj, md_item) else "Missing"

        if not spatial_exist:
            spatial_md_attr = "Dataset Doesn't Exist"
        else:
            if hasattr(spatial_md_obj, md_item):
                spatial_md_attr = _formatMdItems(getattr(spatial_md_obj, md_item),md_item, text_type) if getattr(spatial_md_obj, md_item) else "Missing"
  
        if not service_exist:
            service_md_attr = "Dataset Doesn't Exist"
        else:
            if hasattr(service_obj, SERVICE_ITEM_LOOKUP[md_item]):
                service_md_attr = _formatMdItems(getattr(service_obj, SERVICE_ITEM_LOOKUP[md_item]),SERVICE_ITEM_LOOKUP[md_item], text_type) if getattr(service_obj, SERVICE_ITEM_LOOKUP[md_item]) else "Missing"



        out_dictionary[f"Master - {md_item}"] = master_md_attr
        out_dictionary[f"Spatial - {md_item}"] = spatial_md_attr
        out_dictionary[f"Service - {md_item}"] = service_md_attr

        
        master_md_item_check = _cleanCheckText(master_md_attr) if master_md_attr else None
        spatial_md_item_check = _cleanCheckText(spatial_md_attr) if spatial_md_attr else None
        service_md_item_check = _cleanCheckText(service_md_attr) if service_md_attr else None


        check_list = [i for i in [master_md_item_check, spatial_md_item_check,service_md_item_check] if i is not None and i != "dataset doesn't exist"]
        check_set = set(check_list)
        result = len(check_set)

        if result == 1 and [i for i in check_list][0] == "missing":
            check_bool = False
            
        elif result <= 1:
            check_bool = True

        else:
            check_bool = False

        out_dictionary[f"{md_item} - Match"] = check_bool

    return out_dictionary



        
def _formatMdItems(text:str, md_item:str, text_type:str)->str:
    """
    Using ReGex to strip HTML Tags from the Description, Summary, and Access Constraints.
    The Item Tags are also formatted from a string to a list and all spaces are striped and the list is sorted.
    """

    if text and text_type == "Plain" and md_item in ["description", "accessConstraints", "licenseInfo"]:
        bs = BeautifulSoup(text,'html.parser')

        a_tags = bs.find_all('a', href=True)
        for a_tag in a_tags:
            if a_tag.contents:
                a_tag_contents = a_tag.contents[0]

                if isinstance(a_tag.contents[0], Tag):
                    a_tag_contents = a_tag.contents[0].text
                
                if not a_tag_contents:
                    new_content = f"{a_tag['href']}"
                elif a_tag_contents.startswith("http"):
                    new_content = f"{a_tag['href']}"
                else:
                    new_content = f"{a_tag_contents} ({a_tag['href']})"

                a_tag.contents[0].replace_with(new_content)

            else:
                logger.warning(f"!! HTML Description/AccessConstraint/LicenseInfo is missing <a> tag content.")

        clean_text = bs.get_text()
        #clean_text = re.sub(r'<.*?>', '', new_text)
        
        
    elif text and md_item == 'tags':
        if type(text) == str:
            tag_list = text.split(",")

        if type(text) == list:
            tag_list = text

        cleaned_list = [t.strip() for t in tag_list]

        sorted_list = sorted(cleaned_list)

        clean_text = ",".join(sorted_list)

        
    else:
        clean_text = text


    return clean_text
    


            
def getCatalogRows(catalog_df:DataFrame, gdb_names:list):
    if gdb_names:
        return [r for r in catalog_df.iterrows() if r[1]["Spatial GDB"] in gdb_names]
    else:
        return catalog_df.iterrows()
    

##################################################################################################################################################################
## Main
def main(gis_conn:GIS, gdb_directory:Path, catalog_path:Path, output_excel:Path, text_type:str, spatial_gdb_names:list=None)->None:

    logger.info(f"Starting: {datetime.datetime.now()}")
    logger.info(f"Run by: {os.getlogin()}")
    logger.info(f"Run on: {datetime.datetime.now().strftime('%d/%m/%Y, %H:%M:%S')}")
    logger.info(f"Spatial_GDBs: {spatial_gdb_names}")
    logger.info(f"Output Excel Path: {output_excel}")
    logger.info(f"Output Text Type: {text_type}")
    


    catalog_df = pd.read_excel(io=catalog_path, sheet_name="Inventory",header=0,names=None)

    catalog_rows = getCatalogRows(catalog_df, spatial_gdb_names)

    logger.info(f"Checking Metadata...")


    for index, row in catalog_rows:
        row_obj = DataCatalogRow(row, index, gdb_directory, master_gdb_name, gis_conn)

        logger.info(row_obj.table_name)
        arcpy.AddMessage(row_obj.table_name)

        out_dictionary= getMetadata(row_obj=row_obj, md_items=list(SERVICE_ITEM_LOOKUP.keys()), text_type=text_type)

        DF_DICTIONARY[row_obj.table_name] = out_dictionary
        
 
    

    logger.info(f"Creating DataFrame... {datetime.datetime.now()}")
    df = pd.DataFrame.from_dict(DF_DICTIONARY, "index", columns=DF_COLUMNS)


    ## I sort the Pandas DataFrame for ease of reading. This can be commented out to keep the original order of the Data Catalog.
    df = df.sort_index(axis=0)

    logger.info(f"Exporting DataFrame...")
    df.to_excel(excel_writer=output_excel, sheet_name="MetadataCompare",columns=DF_COLUMNS,index_label="Table Name")

    ########## Excel Workbook ##########
    ## This section deals with the formating of the Excel Workbook
    logger.info(f"Formatting Excel Report...{datetime.datetime.now()}")

    wb = openpyxl.load_workbook(output_excel) ## Creates the Workbook Object

    sheet_name = wb.sheetnames[0] ## Creates a list of strings of the sheet names in the workbook

    ws = wb[sheet_name] ## Creates the Work Sheet object

    ## this conditional checks if the table already exists. If the table exists it is deleted and a new table is created.
    if sheet_name in ws.tables:
        del ws.tables[sheet_name]

    table_dimensions = ws.calculate_dimension() ## grabs the dimensions. these are used when creating the Table Object

    tbl = Table(displayName="MetadataCompare", ref=table_dimensions) ## Creates the Table object using the grabbed dimensions

    style = TableStyleInfo(name=f"TableStyleMedium{2}",showFirstColumn=True, showLastColumn=False, showRowStripes=True, showColumnStripes=False) ## Formats the table. This is where the count comes in.

    tbl.tableStyleInfo = style ## Applies the style info to the Table Object
    ws.add_table(tbl) ## Adds the Table Object to the Worksheet object

    exist_b_cols = ["B", "C", "D"]

    title_cols = ["E", "F","G"]
    title_b_col="H"

    description_cols = [ "I", "J", "K"]
    description_b_col = "L"

    summary_cols = ["M", "N", "O"]
    summary_b_col = "P"

    tags_cols = ["Q", "R", "S"]
    tags_b_col = "T"

    credits_cols =  ["U", "V", "W"]
    credits_b_col = "X"

    access_cols = ["Y", "Z","AA"]
    access_b_col = "AB"


    ## the below deals with the Column Width and Wrapping
    ws.column_dimensions["A"].width = 35

    for i in (exist_b_cols + [title_b_col, description_b_col, summary_b_col, tags_b_col, credits_b_col, access_b_col]):
        ws.column_dimensions[i].width = 20

    for i in list(itertools.chain(title_cols, description_cols, summary_cols, tags_cols, credits_cols, access_cols)):
        ws.column_dimensions[i].width = 75

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)




    wb.save(output_excel)
    del wb

    logger.info(f"Ending: {datetime.datetime.now()}")
